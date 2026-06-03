# -*- coding: utf-8 -*-
"""
convert.py - ExcelのQ&AリストをJSONに変換するスクリプト
=========================================================

【使い方（総務担当者向け）】
  1. data/faq_data.xlsx にQ&Aを追記・編集する
  2. ターミナル（コマンドプロンプト）で以下を実行する:
       python convert.py
  3. web/faq.json が自動生成（更新）される
  4. GitHubにプッシュすれば、ウェブサイトに自動反映される

【Excelの形式】
  ・1行目: ヘッダー行（カテゴリ, 質問, 回答）
  ・2行目以降: Q&Aデータ
  ・回答欄に <br> と入力すると、サイト上で改行として表示される
    例: 詳細は下記の通りです<br>・月曜日は窓口休業<br>・火曜〜金曜は9時〜17時

【インストール（初回のみ）】
  pip install openpyxl
"""

import json
import html
import sys
from pathlib import Path
from datetime import datetime

# openpyxlのインポートチェック
try:
    import openpyxl
except ImportError:
    print("=" * 50)
    print("エラー: openpyxl がインストールされていません")
    print("以下のコマンドを実行してからやり直してください:")
    print("  pip install openpyxl")
    print("=" * 50)
    sys.exit(1)


# ============================================================
# ★ 設定エリア（必要に応じて変更してください）
# ============================================================

# Excelファイルのパス（data/フォルダに置いてください）
EXCEL_PATH = "data/faq_data.xlsx"

# 出力先JSONファイルのパス（変更不要）
OUTPUT_PATH = "web/faq.json"

# Excelのシート名（Noneにすると先頭シートを自動選択）
SHEET_NAME = None

# Excelの列番号（A列=1, B列=2, C列=3）
COL_CATEGORY = 1  # A列: カテゴリ
COL_QUESTION  = 2  # B列: 質問
COL_ANSWER    = 3  # C列: 回答

# ヘッダー行番号（この行はスキップされます）
HEADER_ROW = 1

# ============================================================


def sanitize_answer(text: str) -> str:
    """
    回答テキストをサニタイズする。
    ・HTMLの特殊文字はエスケープ（セキュリティ対策）
    ・<br> タグだけは改行として保持
    ・Excelセル内の改行（Alt+Enter）も <br> に変換
    """
    # まず全てのHTMLをエスケープ（< > & などを無害化）
    safe = html.escape(text, quote=False)

    # <br> タグだけ元に戻す（大文字・小文字どちらも対応）
    safe = safe.replace("&lt;br&gt;", "<br>")
    safe = safe.replace("&lt;BR&gt;", "<br>")
    safe = safe.replace("&lt;Br&gt;", "<br>")

    # Excelのセル内改行（Alt+Enter）も <br> に変換
    safe = safe.replace("\n", "<br>")
    safe = safe.replace("\r", "")

    return safe


def convert_excel_to_json() -> bool:
    """
    Excelファイルを読み込み、FAQデータのJSONを生成する。
    戻り値: True=成功, False=失敗
    """

    # --- Excelファイルの存在確認 ---
    excel_path = Path(EXCEL_PATH)
    if not excel_path.exists():
        print(f"エラー: '{EXCEL_PATH}' が見つかりません")
        print(f"  → '{EXCEL_PATH}' にExcelファイルを置いてください")
        return False

    # --- 出力フォルダの作成 ---
    Path(OUTPUT_PATH).parent.mkdir(parents=True, exist_ok=True)

    # --- Excelを読み込む ---
    print(f"読み込み中: {EXCEL_PATH}")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"エラー: Excelファイルを開けませんでした: {e}")
        return False

    # シートを選択（SHEET_NAMEがNoneなら先頭シート）
    if SHEET_NAME:
        if SHEET_NAME not in wb.sheetnames:
            print(f"エラー: シート '{SHEET_NAME}' が見つかりません")
            print(f"  利用可能なシート: {wb.sheetnames}")
            return False
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active

    print(f"シート名: {ws.title}")

    # --- 行を読み込んでFAQリストを作成 ---
    faq_list = []
    skip_count = 0

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        category = row[COL_CATEGORY - 1] if len(row) >= COL_CATEGORY else None
        question  = row[COL_QUESTION  - 1] if len(row) >= COL_QUESTION  else None
        answer    = row[COL_ANSWER    - 1] if len(row) >= COL_ANSWER    else None

        # 質問または回答が空の行はスキップ
        if not question or not answer:
            skip_count += 1
            continue

        # 文字列に変換してトリミング
        category_str = str(category).strip() if category else "その他"
        question_str  = str(question).strip()
        answer_str    = sanitize_answer(str(answer).strip())

        faq_list.append({
            "id":       len(faq_list) + 1,
            "category": category_str,
            "question": question_str,
            "answer":   answer_str,
        })

    # --- JSONに書き出す ---
    output_data = {
        "_comment":   "このファイルは convert.py によって自動生成されます。直接編集しないでください。",
        "updated_at": datetime.now().strftime("%Y年%m月%d日 %H:%M"),
        "total":      len(faq_list),
        "faqs":       faq_list,
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)

    # --- 結果を報告 ---
    print()
    print("=" * 50)
    print(f"  変換完了！")
    print(f"  FAQデータ数  : {len(faq_list)} 件")
    if skip_count:
        print(f"  スキップ行   : {skip_count} 行（質問または回答が空）")
    print(f"  出力ファイル : {OUTPUT_PATH}")
    print("=" * 50)
    print()
    print("次のステップ:")
    print("  ウェブサイトに反映するには、GitHubにプッシュしてください。")
    print()
    return True


if __name__ == "__main__":
    convert_excel_to_json()
