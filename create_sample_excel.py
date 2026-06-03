# -*- coding: utf-8 -*-
"""
create_sample_excel.py - サンプルExcelファイルを生成するスクリプト
=================================================================
既存のExcelファイルがある場合は不要です。
このスクリプトは、Excelファイルがない場合のテスト用です。

実行方法:
  python create_sample_excel.py

実行すると data/faq_data.xlsx が生成されます。
その後、convert.py を実行してJSONを生成してください。
"""

from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("pip install openpyxl を実行してください")
    exit(1)


SAMPLE_DATA = [
    # (カテゴリ, 質問, 回答)
    ("勤怠・休暇", "有給休暇の申請方法を教えてください",
     "社内ポータルの「勤怠管理」から申請できます。\n取得希望日の3営業日前までに申請してください。\n緊急の場合は直属の上長に口頭で連絡した上で、事後申請も可能です。"),

    ("勤怠・休暇", "半日有給は取得できますか？",
     "はい、取得可能です。<br>午前半休: 9:00〜13:00<br>午後半休: 13:00〜18:00<br>申請は社内ポータルから行ってください。"),

    ("勤怠・休暇", "残業申請はどうすればよいですか？",
     "事前に上長の承認が必要です。<br>やむを得ない場合は当日中に申請してください。"),

    ("経費・精算", "交通費の精算方法を教えてください",
     "社内ポータルの「経費精算」から申請してください。<br>領収書またはICカード利用履歴を添付してください。<br>締め切りは毎月末日で、翌月25日に支給されます。"),

    ("経費・精算", "接待交際費の上限はいくらですか？",
     "1回あたり5,000円（税込）が上限です。<br>超える場合は事前に部長の承認が必要です。"),

    ("備品・設備", "オフィス備品を注文したいのですが",
     "総務部へメールでご依頼ください。<br>品名、数量、用途、希望納期を記載してください。"),

    ("入退社・異動", "退職手続きのながれを教えてください",
     "① 上長に退職の意志を伝える<br>② 退職届を総務部に提出<br>③ 業務の引き継ぎ<br>④ 貸与品の返却<br>⑤ 退職手続き完了"),

    ("福利厚生", "健康診断はいつ受けられますか？",
     "年1回、毎年6〜7月に実施しています。<br>詳細は5月末に全社メールでご案内します。"),
]


def create_sample():
    Path("data").mkdir(exist_ok=True)
    output_path = "data/faq_data.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FAQ"

    # ----- ヘッダー行のスタイル -----
    header_fill = PatternFill(fill_type="solid", fgColor="1a4fa0")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, bottom=thin, top=thin)

    headers = ["カテゴリ", "質問", "回答"]
    col_widths = [18, 40, 60]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 24

    # ----- データ行 -----
    even_fill = PatternFill(fill_type="solid", fgColor="F5F7FB")
    data_font = Font(size=10)

    for row_idx, (cat, q, a) in enumerate(SAMPLE_DATA, start=2):
        fill = even_fill if row_idx % 2 == 0 else None

        for col_idx, value in enumerate([cat, q, a], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 40

    # ----- ウィンドウ枠の固定（ヘッダー行を常に表示） -----
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"サンプルExcelを生成しました: {output_path}")
    print()
    print("次のステップ:")
    print("  python convert.py  ← JSONに変換する")


if __name__ == "__main__":
    create_sample()
